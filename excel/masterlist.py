# Standard library
from pathlib import Path
import os

# Third party modules
import openpyxl
from openpyxl.utils import column_index_from_string
from dotenv import load_dotenv

# Local imports
from excel import ExcelHandler
from masterlist_json import (
    INPUT_COLUMN_MAPPING,
    OUTPUT_COLUMN_MAPPING,
    RANK_CORRECTIONS,
    TENURE_STATUS_CORRECTIONS,
    DEPT_CODE_CORRECTIONS,
    DEPT_CODE,
    env_path
)


class FacultyMasterlistHandler(ExcelHandler):
    """
    Faculty masterlist excel file handler
    """

    # Initialize new workbook where data will be copied
    # Shared between all instances of the class
    new_wb = openpyxl.Workbook()
    new_wb.active.title = 'Faculty List'

    def __init__(self, path):
        super().__init__(path)

        self.INPUT_COLUMN_MAPPING = INPUT_COLUMN_MAPPING
        self.OUTPUT_COLUMN_MAPPING = OUTPUT_COLUMN_MAPPING
        self.RANK_CORRECTIONS = RANK_CORRECTIONS
        self.TENURE_STATUS_CORRECTION = TENURE_STATUS_CORRECTIONS
        self.DEPT_CODE_CORRECTIONS = DEPT_CODE_CORRECTIONS
        self.DEPT_CODE = DEPT_CODE

    def _write_headers(self, sheet):
        """
        Writes dictionary keys in `OUTPUT_COLUMN_MAPPING` as headers on first row in `sheet`.
        """
        for header in self.OUTPUT_COLUMN_MAPPING.keys():
            column = column_index_from_string(self.OUTPUT_COLUMN_MAPPING[header])
            sheet.cell(row=1, column=column).value = header

    def _apply_corrections(self, column, corrections_mapping):
        """
        Replaces values in `column` based on the dictionary `corrections_mapping`.
        """
        for cell in column:
            if cell.value in corrections_mapping:
                cell.value = corrections_mapping[cell.value]

    def _copy_whole_sheet(self, sheet, target_sheet):
        """
        Copies all cells of `sheet` to `target_sheet`
        """
        for i in range(4, sheet.max_row):
            for j in range(1, sheet.max_column):
                target_sheet.cell(row=i - 2, column=j).value = sheet.cell(row=i, column=j).value

    def _selective_copy(self, sheet, target_sheet):
        """
        Goes through the first column in `sheet`, if a value matches a key in `INPUT_COLUMN_MAPPING`, then
        values in `INPUT_COLUMN_MAPPING` will indicate the type of data. `OUTPUT_COLUMN_MAPPING` will indicate
        which column that type of data needs to be copied to.

        For example,

        INPUT... = {"Tenured": {"Tenure_Status": "A", ...}}
        OUTPUT... = {"Tenure_Status": "E", ...}

        If the value in the first column is Tenured, then tenure status is in column A,
        which will be copied to column E in the new sheet.

        Also fills in years and department/division codes.
        """
        dept_code = self.DEPT_CODE_CORRECTIONS.get(sheet.title, sheet.title)  # Default value is sheet title unless a correction exists
        div_code = self.DEPT_CODE.get(dept_code, 'N/A')
        fall_year = sheet['A1'].value.strip()[-4::]  # Extract from the first cell, last 4 chars is a year of the fall semester
        acad_year = f'{fall_year}/{int(fall_year[-2::]) + 1}'  # Derive from fiscal year
        self.fiscal_year = int(fall_year) + 1

        valid_values = INPUT_COLUMN_MAPPING.keys()
        max_target_sheet = target_sheet.max_row
        i = 1

        for cell in sheet['A']:
            tenure_status = cell.value
            if tenure_status in valid_values:
                # Ignore if name is empty
                name_column = column_index_from_string(INPUT_COLUMN_MAPPING[tenure_status]['Name'])
                name_value = sheet.cell(row=cell.row, column=name_column).value
                if name_value:
                    # Go through columns and perform copying
                    for data_type, input_column_letter in INPUT_COLUMN_MAPPING[tenure_status].items():
                        input_column = column_index_from_string(input_column_letter)
                        output_column = column_index_from_string(self.OUTPUT_COLUMN_MAPPING[data_type])
                        target_sheet.cell(row=max_target_sheet + i, column=output_column).value = \
                            sheet.cell(row=cell.row, column=input_column).value

                    # Fill-in extra values
                    target_sheet.cell(row=max_target_sheet + i,
                                      column=column_index_from_string(self.OUTPUT_COLUMN_MAPPING['Fiscal_Year'])).value = self.fiscal_year
                    target_sheet.cell(row=max_target_sheet + i,
                                      column=column_index_from_string(self.OUTPUT_COLUMN_MAPPING['Academic_Year'])).value = acad_year
                    target_sheet.cell(row=max_target_sheet + i,
                                      column=column_index_from_string(self.OUTPUT_COLUMN_MAPPING['Division_Code'])).value = div_code
                    target_sheet.cell(row=max_target_sheet + i,
                                      column=column_index_from_string(self.OUTPUT_COLUMN_MAPPING['Department_Code'])).value = dept_code

                    i += 1

    def clean(self):
        """
        Performs steps in the cleaning workflow:
            1. Write headers to the new workbook
            2. Apply corrections to tenure status and rank
            3. Copy data from original to new workbook
        """
        self._write_headers(self.new_wb.active)

        for sheet_name in self.wb.sheetnames:
            self._apply_corrections(self.wb[sheet_name]['A'], self.TENURE_STATUS_CORRECTION)
            self._apply_corrections(self.wb[sheet_name]['B'], self.RANK_CORRECTIONS)
            self._selective_copy(sheet=self.wb[sheet_name], target_sheet=self.new_wb.active)

        output_folder = Path(os.getenv('OUTPUT_FOLDER_PATH'))
        self.new_wb.save(output_folder / f'FY{self.fiscal_year}_FTE.xlsx')


if __name__ == "__main__":

    load_dotenv(dotenv_path=env_path)

    PATHS = [
        os.getenv('ARTS_SPS_PATH'),
        os.getenv('HUM_PATH'),
        os.getenv('NS_PATH'),
        os.getenv('SS_PATH'),
    ]

    for wb_path in PATHS:
        division = FacultyMasterlistHandler(wb_path)
        division.clean()
