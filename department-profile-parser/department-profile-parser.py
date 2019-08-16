# === IMPORTS ===

# Standard library imports
import os
import json
import sys

# Third party imports
from dotenv import load_dotenv
from pathlib import Path
import openpyxl

# === FUNCTION DEFINITIONS ===

def progress_bar(n, n_max):
    sys.stdout.write('\r')
    j = (n + 1) / n_max
    sys.stdout.write("[%-20s] %d%%" % ('='*int(j*20), j*100))
    sys.stdout.flush()

def load_json(json_path):
    with open(json_path, 'r') as read_file:
        json_data = json.load(read_file)
    return json_data

def copyValues(input_sheet, output_sheet, mapping_array):
    """Copies the values from the profile sheet

    - Takes the input Excel sheet
    - Iterates through columns with years
    - Copies values to the output Excel sheet according to the mapping array

    Items in each mapping array of the json file:
    0: column number in output sheet
    1: row number in the department profile excel sheet from the current column
    2: how many columns to offset from the year column
       (each year takes up two columns: first the number of enrollments, then the number of classes)

    Args:
        input_sheet: where the values are taken FROM
        output_sheet: where the values are going TO
        mapping_array: mapping between the two sheets (where a particular cell from the input sheet should go in the output sheet)
    
    Returns:
        Nothing

    Mapping example: 
        [65, 176, 0] means:
        1. Go to row 176 of the current column without offsetting
        2. Copy to the new row of column 65 of the output sheet
    """

    new_row = output_sheet.max_row + 1
    column_with_year = range(6,35,2) # Each year takes up two columns

    for year in column_with_year:
        # Iterate through all values
        for new_column, old_row, offset in mapping_array:
            input_value = input_sheet.cell(row=old_row, column=year+offset).value
           
            # If copying selectivity/yield values, then keep blank cells blank
            # Otherwise replace blanks with zeroes
            logicalSelectivityYield = (new_column in (10, 16, 23, 11, 17, 24, 37, 38)) \
                                      and (old_row in (29, 35, 33, 30, 36, 44, 30, 42, 50, 31, 43, 51, 36, 37))
            logicalBlank = (input_value is None) or (input_value is '') or (input_value == 'n/a')
            
            if (logicalSelectivityYield and logicalBlank) or (not logicalBlank):
                output_sheet.cell(row=new_row, column=new_column).value = input_value
            else:
                output_sheet.cell(row=new_row, column=new_column).value = 0

            # Department name has to be copied separately because it's in the same place every time
            output_sheet.cell(row=new_row, column=2).value = input_sheet.cell(row=4, column=1).value 
        
        # Each year is on a new row
        new_row += 1

# ============================

# Base directory of the script file
basedir = Path(__file__).parent.resolve()

# Load environment variables
env_path = basedir / '.env'
load_dotenv(dotenv_path=env_path)

# Load mapping arrays from the mapping json
mapping_path = basedir / 'mapping.json'
mapping_data = load_json(mapping_path)

# Load sheet names of each department with corresponding mapping arrays
departments_mapping_path = basedir / 'departments.json'
departments = load_json(departments_mapping_path)

# Load the Excel files
input_file_path = Path(os.getenv('INPUT_FILE_PATH'))
input_file = openpyxl.load_workbook(input_file_path, data_only=True) # data_only=True will make sure that that a cell value does not return the formula text

output_file_path = Path(os.getenv('OUTPUT_FILE_PATH'))
output_file = openpyxl.load_workbook(output_file_path)

# Iterate through every department in the input file based on the sheet named loaded from the json 
n = 0
n_max = len(departments) 
for department_sheet_name, department_mapping_list in departments.items():
    input_sheet = input_file[department_sheet_name]
    
    # Iterate through each mapping array
    for mapping_array_name in department_mapping_list:
        output_sheet = output_file[mapping_data[mapping_array_name]['output_sheet']]
        mapping_array = mapping_data[mapping_array_name]['mapping']
        copyValues(input_sheet, output_sheet, mapping_array)
        
    n = n + 1
    progress_bar(n, n_max)
 
# Save output file 
output_file.save(output_file_path)
